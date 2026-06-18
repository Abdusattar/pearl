"""
Экспорт списка детей для заполнения групп Мунарой.
Запуск: python scripts/export_students_for_groups.py
Выходной файл: data/дети_группы.xlsx
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.database import SessionLocal
from app.models import Student

OUTPUT = Path("data/дети_группы.xlsx")


def main():
    db = SessionLocal()
    students = db.query(Student).filter(Student.status == "active").order_by(Student.pin).all()
    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Дети — группы"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2E75B6")
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    center   = Alignment(horizontal="center", vertical="center")
    left     = Alignment(horizontal="left", vertical="center")
    thin     = Side(style="thin", color="BFBFBF")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["№", "PIN", "ФИО ребёнка", "Локация (Сокулук / Кожомкул)", "Группа"]
    widths  = [5, 7, 32, 30, 20]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, col, h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    for i, s in enumerate(students, 1):
        fill = alt_fill if i % 2 == 0 else None
        row_data = [i, s.pin, s.name, "", ""]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(i + 1, col, val)
            cell.border    = border
            cell.alignment = center if col != 3 else left
            if fill:
                cell.fill = fill
        ws.row_dimensions[i + 1].height = 18

    ws.freeze_panes = "A2"
    wb.save(OUTPUT)
    print(f"Готово: {OUTPUT} — {len(students)} детей")


if __name__ == "__main__":
    main()
