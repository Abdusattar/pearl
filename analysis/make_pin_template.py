"""Создаёт Excel-шаблон списка детей с автогенерацией PIN-кода."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "pin_list_template.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Список детей садика"

# Стили
header_font  = Font(bold=True, color="FFFFFF", size=11)
header_fill  = PatternFill("solid", fgColor="2E75B6")
year_font    = Font(bold=True, color="FFFFFF", size=10)
year_fill    = PatternFill("solid", fgColor="5B9BD5")
center       = Alignment(horizontal="center", vertical="center")
thin         = Side(style="thin", color="BFBFBF")
border       = Border(left=thin, right=thin, top=thin, bottom=thin)

# Ширины колонок
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 12

# Заголовок
headers = ["№", "ФИО ребёнка", "Дата рождения\n(ДД.ММ.ГГГГ)", "PIN-код"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[1].height = 32

# Разделитель года
def add_year_row(row, year):
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws.cell(row=row, column=1, value=f"── {year} год рождения ──")
    cell.font = year_font
    cell.fill = year_fill
    cell.alignment = center
    ws.row_dimensions[row].height = 18

# Пример данных
examples = [
    (2021, [
        (1, "Асанов Изат",     "02.06.2021"),
        (2, "Пример Ребёнок",  "15.09.2021"),
    ]),
    (2022, [
        (3, "Аманова Имана",   "03.05.2022"),
        (4, "Исабекова Бурул", "03.05.2022"),
        (5, "Пример Ребёнок",  "20.11.2022"),
    ]),
    (2023, [
        (6, "Пример Ребёнок",  "10.01.2023"),
    ]),
]

data_rows = []  # список строк Excel где есть данные (для формулы)
row = 2
for year, children in examples:
    add_year_row(row, year)
    row += 1
    for num, name, dob in children:
        data_rows.append(row)
        ws.cell(row=row, column=1, value=num).alignment = center
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=dob).alignment = center
        # PIN формула: DD + MM + YY (последние 2 цифры) + порядковый номер при совпадении дат
        ws.cell(row=row, column=4,
                value=f'=LEFT(C{row},2)&MID(C{row},4,2)&MID(C{row},9,2)&COUNTIF($C$2:C{row},C{row})'
                ).alignment = center
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
        row += 1

# Пустые строки для заполнения (ещё 40 строк с формулой)
for i in range(40):
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        if col == 4:
            cell.value = f'=IF(C{row}="","",LEFT(C{row},2)&MID(C{row},4,2)&MID(C{row},9,2)&COUNTIF($C$2:C{row},C{row}))'
            cell.alignment = center
    row += 1

# Заморозить шапку
ws.freeze_panes = "A2"

wb.save(OUT)
print(f"Готово: {OUT}")
