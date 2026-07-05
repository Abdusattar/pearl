import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

wb_in = openpyxl.load_workbook('ИНН.xlsx')
ws_in = wb_in.active

children = []
for row in range(1, ws_in.max_row + 1):
    name   = ws_in.cell(row, 1).value
    inn    = str(ws_in.cell(row, 2).value or '')
    amount = ws_in.cell(row, 3).value
    status = ws_in.cell(row, 4).value
    if not name or len(inn) != 14:
        continue
    dd, mm, yyyy = inn[1:3], inn[3:5], inn[5:9]
    yy = yyyy[2:]
    dob_key = f'{dd}{mm}{yy}'
    children.append({'name': name, 'inn': inn, 'dd': dd, 'mm': mm,
                     'yyyy': yyyy, 'yy': yy, 'dob_key': dob_key,
                     'amount': amount, 'status': status})

# PIN: DDMMYY + порядковый (стабильный, по алфавиту внутри даты)
groups = defaultdict(list)
for c in children:
    groups[c['dob_key']].append(c)

for key, group in groups.items():
    group.sort(key=lambda x: x['name'])
    for i, c in enumerate(group, 1):
        c['seq'] = i
        c['pin'] = c['dob_key'] + str(i)

wb_out = openpyxl.Workbook()
ws = wb_out.active
ws.title = 'PIN-список садик'

hdr_font  = Font(bold=True, color='FFFFFF', size=11)
hdr_fill  = PatternFill('solid', fgColor='2E75B6')
alt_fill  = PatternFill('solid', fgColor='EBF3FB')
exit_fill = PatternFill('solid', fgColor='F2DCDB')
center    = Alignment(horizontal='center', vertical='center')
thin      = Side(style='thin', color='BFBFBF')
border    = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ['№', 'ФИО ребёнка', 'Дата рождения', 'PIN-код', 'ИНН', 'Статус']
widths  = [5, 30, 16, 12, 18, 12]
for col, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(1, col, h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = center
    cell.border = border
    ws.column_dimensions[cell.column_letter].width = w
ws.row_dimensions[1].height = 24

active = [c for c in children if not c['status']]
exited = [c for c in children if c['status'] == 'выбыл']

for i, c in enumerate(active + exited, 1):
    is_exit = bool(c['status'])
    row_fill = exit_fill if is_exit else (alt_fill if i % 2 == 0 else None)
    dob_str = f"{c['dd']}.{c['mm']}.{c['yyyy']}"
    vals = [i, c['name'], dob_str, c['pin'], c['inn'], c['status'] or 'активный']
    for col, val in enumerate(vals, 1):
        cell = ws.cell(i + 1, col, val)
        cell.border = border
        cell.alignment = center if col != 2 else Alignment(vertical='center')
        if row_fill:
            cell.fill = row_fill

ws.freeze_panes = 'A2'
wb_out.save('PIN_список_садик.xlsx')

print(f'Готово: PIN_список_садик.xlsx')
print(f'Активных: {len(active)} | Выбыло: {len(exited)} | Всего: {len(children)}')
print()
print('Первые 10 активных:')
for c in active[:10]:
    print(f"  {c['name']} | {c['dd']}.{c['mm']}.{c['yyyy']} → PIN: {c['pin']}")
print()
dups = {k: v for k, v in groups.items() if len(v) > 1}
if dups:
    print(f'Дублирующиеся даты ({len(dups)} случаев):')
    for key, group in dups.items():
        print(f"  {key}: " + ', '.join(f"{c['name']} → {c['pin']}" for c in group))
